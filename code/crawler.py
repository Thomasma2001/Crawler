#载入必要package
import sys
import time
from bs4 import BeautifulSoup
import re
import urllib.request as urllib
import urllib.error
import openpyxl

#定义一个方法, 获得页面全部内容
def askURL(url):
    headers={"User-Agent" : "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE"}#缺这句原始文件不能运行
    request = urllib.request.Request(url,headers=headers)#发送请求
    try:
        response = urllib.request.urlopen(request)#取得响应
        html= response.read()#获取网页内容
        # print (html)
    except urllib.error.URLError as e:
        if hasattr(e,"code"):
            print (e.code)
        if hasattr(e,"reason"):
            print (e.reason)
    return html

#获取相关内容
#Warning: 解析页面的代码和页面结构强相关, 当页面html结构发生变化的时候必须同步升级, 否则当场作废

def getData(baseurl):
    findLink=re.compile(r'<a href="(.*?)">')#找到影片详情链接
    findImgSrc=re.compile(r'<img.*src="(.*?)"',re.S)#找到影片图片
    findTitle=re.compile(r'<span class="title">(.*)</span>')#找到片名
    #找到评分
    findRating=re.compile(r'<span class="rating_num" property="v:average">(.*)</span>')
    #找到评价人数
    findJudge=re.compile(r'<span>(\d*)人评价</span>')
    #找到概况
    findInq=re.compile(r'<span class="inq">(.*)</span>')
    #找到影片相关内容：导演，主演，年份，地区，类别
    findBd=re.compile(r'<p class="">(.*?)</p>',re.S)
    #去掉无关内容
    remove=re.compile(r'                            |\n|</br>|\.*')
    datalist=[]

    for i in range(0,10):
        url=baseurl+str(i*25)
        html=askURL(url)
        soup = BeautifulSoup(html, "html.parser")
        for item in soup.find_all('div',class_='item'):#找到每一个影片项
            data=[]
            item=str(item)#转换成字符串
            # 影片详情链接
            link=re.findall(findLink,item)[0]
            data.append(link)#添加详情链接  
            imgSrc=re.findall(findImgSrc,item)[0]
            data.append(imgSrc)#添加图片链接
            titles=re.findall(findTitle,item)
            #片名可能只有一个中文名，没有外国名
            if(len(titles)==2):
                ctitle=titles[0]
                data.append(ctitle)#添加中文片名
                otitle=titles[1].replace("/","")#去掉无关符号
                data.append(otitle)#添加外国片名
            else:
                data.append(titles[0])#添加中文片名
                data.append(' ')#留空

            rating=re.findall(findRating,item)[0]
            data.append(rating)#添加评分
            judgeNum=re.findall(findJudge,item)[0]
            data.append(judgeNum)#添加评论人数
            inq=re.findall(findInq,item)
            
            #可能没有概况
            if len(inq)!=0:
                inq=inq[0].replace("。","")#去掉句号
                data.append(inq)#添加概况
            else:
                data.append(' ')#留空
            bd=re.findall(findBd,item)[0]
            bd=re.sub(remove,"",bd)
            bd=re.sub('<br(\s+)?\/?>(\s+)?'," ",bd) #去掉<br >
            bd=re.sub('/', " ",bd)#替换/
            data.append(bd.strip())
            datalist.append(data)
            
    time.sleep(5)
    return datalist

#将相关数据写入excel中

def saveData(datalist,savepath):
    book=openpyxl.Workbook()
    sheet = book.create_sheet("豆瓣电影Top250")
    #sheet=book.add_sheet('豆瓣电影Top250',cell_overwrite_ok=True)
    col=('电影详情链接','图片链接','影片中文名','影片外国名', '评分','评价数','概况','相关信息')

    sheet.append(col)  #添加列头

    
    for i in range(0,250):
        data=datalist[i]
        for j in range(0,8):
            sheet.cell(row = (i+2),column = (j+1),value = data[j]) 

#openpyxl中的单元格计数从1开始, 加上第一行是列头, 要多跳一个
#openpyxl中的计数和Excel内的计数方式一致, 但和常规编程从0开始的方式相左

    book.save(savepath) #保存
    
def main():

    print ("开始爬取......")
    baseurl='https://movie.douban.com/top250?start='
    datalist=getData(baseurl)
    savapath=u'/devnet/paddle/web-crawler/豆瓣电影Top250.xlsx'
    saveData(datalist,savapath)
    
    
main()

print ("爬取完成，请查看.xlsx文件")

from pandas import Series,DataFrame
import pandas as pd
#打印一些样例数据观察

from openpyxl import load_workbook
wb = load_workbook(filename = '/devnet/paddle/web-crawler/豆瓣电影Top250.xlsx')
sheet_ranges = wb['豆瓣电影Top250']

#将第一列的前几行打印出来
for i in range (1,8):
    print(sheet_ranges['A'+i.__str__()].value)
    

#如果没有问题, 就将该工作表转为dataframe格式






